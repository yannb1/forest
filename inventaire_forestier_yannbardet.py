#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Dec 16 02:19:57 2017

@author: yannbardet
"""

# =============================================================================
# projet yann bardet INVENTAIRE PAR PLACETTE (rayon constant)
# =============================================================================

#divers imports 
import math
import xlsxwriter
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os

#variables globales
rayon=0
fiche='' #numéro de la fiche du fichier de données(xlsx): choix: 1431, 1350, 1384, 1412
path='' #path absolu du fichier .py

#Fonction permettant à l'utilisateur d'entrer le path du fichier contenant le
#skript python actuellement utilisé
def entrer_path():
    global path #appeler la variable globale pour pouvoir la modifier
    print("Bienvenue. Veuillez entrer les données qui vous sont demandées")
    path=input("Entrer la path du dossier dans lequel se trouve le fichier"
                  " \"inventaire_forestier_yannbardet.py\". \nEXEMPLE: "
                  "/Users/yannb/Desktop/projet_python_yannbardet.  \nATTENTION! le"
                  " fichier Excel \"data_inventaire\" doit également se trouver "
                  "dans ce fichier!  :")
    entrer_no_fiche()
    
#fonction permettant à l'utilisateur de choisir quelle placette traiter
def entrer_no_fiche():
    global fiche
    fiche=input("Entrer le numéro de la placette à traiter. "
                "Ce numéro correspond au num. de fiche des données excel."
                "(choix:1431, 1350, 1384, 1412) :")
    entrer_rayon()
    
#permet à l'utilisateur de rentrer le rayon correspondant
def entrer_rayon():
    global rayon
    rayon=float(input("Entrer le rayon de la placette à analyser selon la pente"
                "(dans notre cas le rayon est de 12.69[m]):"))
    calcul()
    

#lance les calculs, création des plot et écriture du nouveau fichier excel
def calcul():
    print("Les calculs sont effectués, les plot dessinés et le tout est exporté "
          "dans un nouveau fichier excel (résultats inventaire) se trouvant dans"
          "le même dossier que le fichier excel contenant les données ainsi que "
          "le fichier .py .")
        
    #----------------------------------------------------
    #importer donnée externe
    
    #lieu ou python va chercher et enregistrer les information (excel)
    #à partir du path entré par l'utilisateur
    os.chdir(path)
    #ouvrir le fichier xlsx, à partir du path précédamment défini
    wb1 = openpyxl.load_workbook('data_inventaire.xlsx') 
      
    #défini quelle page du fichier excel source sera utilisée à partir de la variable
    # "fiche" définie au début
    sheet=wb1.get_sheet_by_name(fiche)
    
    #----------------------------------------------------
    #mettre les données dans une matrice
    # création d'une matrice nulle de la taille du tableau excel 
    MyData =  [[0 for e in range(2,sheet.max_column)] for e in range(0,sheet.max_row-1)]
    nbre_ligne=(len(MyData) ) #nnombre de ligne: pour les "range" des boucles à venir
    
    #ajouter chaque colonne du fichier excel à la matrice
    for row in range(2,sheet.max_row+1,1):
        MyData[(row-2)][0]=sheet['A'+ str(row)].value 
        MyData[(row-2)][1]=sheet['B'+ str(row)].value
        MyData[(row-2)][2]=sheet['C'+ str(row)].value
        MyData[(row-2)][3]=sheet['D'+ str(row)].value  
        MyData[(row-2)][4]=sheet['E'+ str(row)].value
    #row-2=ligne 0 de la matrice: les premières valeurs sont dans la ligne 2 du fichier excel
    #ligne 2-2 = ligne 0 matrice
    
    #----------------------------------------------------
    #liste vide pour résineux et feuillus
    resineux=[]
    feuillus=[]
    #sortir les lignes de la matrice "mydata" qui concernent un résineux, et les 
    #ajouter à la liste(matrice) "résieux"
    #idem pour les feuillus
    for i in range(0,nbre_ligne):
        if MyData[i][3] in["TA","FI","FO","LA","EE","DO","WF","SF","DF","ND"]:
            resineux.append(MyData[i])   
        else:
            feuillus.append(MyData[i])

    #----------------------------------------------------
    #calcul des coordonnées de chaque arbre
    #initialisation vecteur coordonnées x et y
    xcoord=[]
    ycoord=[]  
    #initialisation compteur     
    j=0
    #boucle pour le calcul des coordonnées à partir des azimuts et des distances
    while j < nbre_ligne:
        #coordonnée x= sin(angle(rad))*distance
        xcoord.append((math.sin(MyData[j][1]*math.pi/200))*MyData[j][2]) #math.pi/200 : transformation des grad en rad --> math.sin fonctionne avec des rad
        #coordonnée y= cos(angle)*distance
        ycoord.append((math.cos(MyData[j][1]*math.pi/200))*MyData[j][2]) #math.pi/200 : transformation des grad en rad --> math.cos fonctionne avec des rad   
        j+=1
    #----------------------------------------------------    
    #initialisation matrice coordonnées raisieux et feuillus pour leur différenciation dans le plot
    xcoor_resineux=[]
    ycoor_resineux=[]
    xcoor_feuillus=[]
    ycoor_feuillus=[]
    
    #boucle pour séparer coordonnées x et y des feuillus et raisineux
    for i in range(0,nbre_ligne): #pour chaque ligne de MyData
        #si ligne i correspond à un résineux
        if  MyData[i][3] in["TA","FI","FO","LA","EE","DO","WF","SF","DF","ND"]:
            xcoor_resineux.append(xcoord[i])
            ycoor_resineux.append(ycoord[i])
        else: #si feuillus : idem que pour résineux
            xcoor_feuillus.append(xcoord[i])
            ycoor_feuillus.append(ycoord[i])
    
    # =============================================================================
    # Calculs
    # =============================================================================
    #superficie placette 
    surface=math.pi*rayon**2
    #calcul de N: nombre de tige par hectare
    #Feuillus
    N_tiges_feu=int(len(feuillus)*10000/surface) #nombre arbres placette* 1 ha/surface placette= nombre arbre à l'hectare
    #convertion en integer puisque 33,23 arbres n'a pas de sens-->entier 
    #résineux
    N_tiges_res=int(len(resineux)*10000/surface)   
    
    #total
    N_tiges_tot=int(len(MyData)*10000/surface )                    
    
    #----------------------------------------------------
    #Calcul de G: surface terrière [m^2/ha]
    #feuillus
    surf_terr_feu=[] #création liste vide pour receuillir données de la boucle
    for i in range(0,len(feuillus)): #boucle pour chaque ligne de la matrice feuillus
        #G=rayon**2 *pi  /!\ rayon en mètre--> diamètre[cm]/200
        surf_terr_feu.append(((feuillus[i][4]/200)**2)*math.pi)
        
    #addition des surf. terr. de chaque arbre
    #extrapolation pour l'ensembée d'un ha
    surf_terr_feu_tot=round(sum(surf_terr_feu)*10000/surface,2) 
    #round permet d'arrondir la valeur obtenue, dans ce cas à 2 chiffres après la virgule.
    
    #idem pour résnineux
    surf_terr_res=[]    
    for i in range(0,len(resineux)):
        #G=rayon**2 *pi  /!\ rayon en mètre--> diamètre[cm]/200
        surf_terr_res.append(((resineux[i][4]/200)**2)*math.pi)   
    
    surf_terr_res_tot=round(sum(surf_terr_res)*10000/surface,2)
    
    #surface terrière totale
    surf_terr_tot=surf_terr_feu_tot+surf_terr_res_tot
    #----------------------------------------------------
                                                                                                     
    #calcul de dg (diamètre de la tige à la surface terrière moyenne)
    #feuillus
    #boucle qui fait la somme des diamètres**2
    diam_carr_feu=[]
    for i in range(0,len(feuillus)): #boucle pour chaque ligne de la matrice feuillus
        #Diamètre**2
        diam_carr_feu.append(feuillus[i][4]**2)
    #dg=sqrt((somme diamètre**2)/Nfeuillus dans placette)
    dg_feu=round(math.sqrt((sum(diam_carr_feu))/len(feuillus)),2)
    
    #résineux
    #boucle qui fait la somme des diamètres**2
    diam_carr_res=[]
    for i in range(0,len(resineux)): #boucle pour chaque ligne de la matrice resineux
        #Diamètre**2
        diam_carr_res.append(resineux[i][4]**2)
    #dg=sqrt((somme diamètre**2)/Nresineux dans placette)
    dg_res=round(math.sqrt((sum(diam_carr_res))/len(resineux)),2)
        
    #dg feuillus résineux ensembles
    #boucle qui fait la somme des diamètres**2
    diam_carr_tout=[]
    for i in range(0,len(MyData)): #boucle pour chaque ligne de la matrice MyData
        #Diamètre**2
        diam_carr_tout.append(MyData[i][4]**2)
    #dg=sqrt((somme diamètre**2)/Nfeuillus dans placette)
    dg_tout=round(math.sqrt((sum(diam_carr_tout))/len(MyData)),2)
    #----------------------------------------------------    
    
    #calcul de V  : volume sur pied (formule plus ou moins approximative, mais donne une idée)
    #V=(somme(diamètre[cm]**2)) /1000   
    #feuillus
    vol_feu=[]
    for i in range(0,len(feuillus)):  #boucle pour chaque ligne de la matrice
        vol_feu.append((feuillus[i][4]**2)/1000)  
    vol_feu_tot=round(sum(vol_feu)*10000/surface,2)   #somme de la liste   + extrapolation pour 1 ha                                            
    
    #résineux
    vol_res=[]
    for i in range(0,len(resineux)):  #boucle pour chaque ligne de la matrice
        vol_res.append((resineux[i][4]**2)/1000)  
    vol_res_tot=round(sum(vol_res)*10000/surface,2)  #somme de la liste  
    
    #V résnineux et feuillus
    vol_tot=round((sum(vol_res)+sum(vol_feu))*10000/surface,2)
        
      
    
    
    # =============================================================================
    # plots
    # =============================================================================
    
    #pour le plot nombre arbres par classes de diamètre de 9 cm  
    #création des classes
    classe1=[] #12-21 cm
    classe2=[] #22-31 cm
    classe3=[] #32-41 cm
    classe4=[] #42-51 cm
    classe5=[] #52-61 cm
    classe6=[] #62-71 cm
    classe7=[] #72-81 cm
    classe8=[] #82-91 cm 
    
    #boucle pour répartir les différents DHP dans les classes
    for i in range(0,len(MyData)):
        if  12 <= MyData[i][4] < 22  : #si le diamètre se trouve entre 12 et 21
            classe1.append(MyData[i][4])
        elif 22 <= MyData[i][4] < 32 : #si le diamètre se trouve entre 22 et 31
            classe2.append(MyData[i][4])
        elif 32 <= MyData[i][4] < 42 : 
            classe3.append(MyData[i][4])
        elif 42<= MyData[i][4] < 52 : 
            classe4.append(MyData[i][4])
        elif 52 <= MyData[i][4] < 62 : 
            classe5.append(MyData[i][4])
        elif 62 <= MyData[i][4] < 72 : 
            classe6.append(MyData[i][4])
        elif 72 <= MyData[i][4] < 82 : 
            classe7.append(MyData[i][4])
        elif 82 <= MyData[i][4] < 92 : 
            classe8.append(MyData[i][4])
    
    fig = plt.figure() #préparation à l'exportation
    plt.rcdefaults()
    fig, ax = plt.subplots()
    #liste avec le nombre d'arbres par classe =axe x
    nbre_classe=[len(classe1),len(classe2),len(classe3),len(classe4),
                 len(classe5),len(classe6),len(classe6),len(classe8) ]
    #liste no classe pour le plot
    no_classe=[1,2,3,4,5,6,7,8] #=axe y
    # histogramme horizontal
    ax.barh(no_classe, nbre_classe, align='center',
            color='green', ecolor='black')
    
    ax.set_xlabel('nombre d\'arbres')
    ax.set_ylabel('Classes de diamètre ')
    ax.set_title('Nombre d\'arbres par classe de diamètre [cm] \n cl.1: 12-21,' 
                 'cl.2: 22-31, cl.3: 32-41, cl.4: 42-51, \n cl.5: 52-61, cl.6: 62-71, cl.7: 72-81, cl.8: 82-91')
    ax.set_xticks(range(0,max(nbre_classe)+1)) #pour avoir axe des x avec nmbre entier
    
    fig.savefig('hist_class_diam.png') #sauvegarder plot en png sous le nom...
    
    
    #----------------------------------------------------
    #plot représentation placette vue du ciel
    fig1 = plt.figure() #préparation à l'importation
    #tracage du cercle du périmètre
    # theta goes from 0 to 2pi 
    theta = np.linspace(0, 2*np.pi, 100)
    # compute x1 and x2 #
    x1 = rayon*np.cos(theta) #pour tracer le cercle avec le rayon voulu
    x2 = rayon*np.sin(theta)      
    # créer le rayon de la placette =le cercle
    plt.plot(x1, x2,"--",label="rayon placette",
             color=(0.5019607843137255, 0.0, 0.0))
    
    #points arbre résineux, vert foncé, avec indication pour la légende
    plt.plot(xcoor_resineux,ycoor_resineux, 'o', label= "résineux", 
             color="g") 
    #points arbres feuillus
    plt.plot(xcoor_feuillus,ycoor_feuillus, 'o', label= "feuillus", 
             color=(0.4980392156862745, 1.0, 0.0)) 
    #centre placette (0,0): croix rouge
    plt.plot(0,0,"X",label="centre placette",color="r") 
    
    #labels des axes et titre
    plt.xlabel('Distance [m]')
    plt.ylabel('Distance [m]')
    plt.title('Placette numéro %r'%fiche)
    
    plt.axis("equal")#axes x et y à la même échelle
    
    #légende en haut à droite: voulais la mettre en dehors, mais ne s'affiche pas après l'export en png
    plt.legend(loc='upper left')
    
    
    #sauvegarder figure au formal .png
    fig1.savefig('plot_résultats.png') #sauvegarder plot en png sous le nom...
    
    
     
    # =============================================================================
    # Création et écriture du fichier excel pour les résultats   
    # =============================================================================
     
    #création du fichier 
    wb  = xlsxwriter.Workbook('résultats inventaire.xlsx') 
    #créaction de la feuille du fichier avec le nom de la placette
    ws = wb.add_worksheet(fiche)
    
    #nom de la placette
    #écrire dans le nouveau fichier dans la cellule (0,2)
    ws.write(0, 0, 'Placette numéro') 
    ws.write(0, 1, fiche) #fiche=numéro placette--> cellule 0,1
    #----------------------------------------------------  
    #surface
    ws.write(0,4,'surface[m2]:')
    ws.write(0,5,surface)
    #----------------------------------------------------  
    #rayon
    ws.write(0,7, 'rayon[m]')
    ws.write(0,8, rayon)
    #----------------------------------------------------  
    #Ecriture des différents résultats obtenus pour N, G, dg et V
    #N
    ws.write(2, 0, 'Nombre de tiges par ha (N)')
    ws.write(3, 0, 'N feuillus')
    ws.write(3, 1, N_tiges_feu)
    ws.write(4, 0, 'N résineux')
    ws.write(4, 1, N_tiges_res)
    ws.write(5, 0, 'N Feuil. + res')
    ws.write(5, 1, N_tiges_tot)
    #----------------------------------------------------  
    #G
    ws.write(2, 3, 'Surface terrière [m2] par ha (G)')
    ws.write(3, 3, 'G feuillus ')
    ws.write(3, 4, surf_terr_feu_tot)
    ws.write(4, 3, 'G résineux')
    ws.write(4, 4, surf_terr_res_tot)
    ws.write(5, 3, 'G feuil.+res')
    ws.write(5, 4, surf_terr_tot)
    #----------------------------------------------------  
    #dg
    ws.write(2, 6, 'Diamètre[cm] de la tige à la surface terrière moyenne (dg)')
    ws.write(3, 6, 'dg feuillus ')
    ws.write(3, 7, dg_feu)
    ws.write(4, 6, 'dg résineux')
    ws.write(4, 7, dg_res)
    ws.write(5, 6, 'dg feuil.+res')
    ws.write(5, 7, dg_tout)
    #----------------------------------------------------  
    #V
    ws.write(7, 0, 'Volume[m3] sur pied par ha (V)')
    ws.write(8, 0, 'V feuillus')
    ws.write(8, 1, vol_feu_tot)
    ws.write(9, 0, 'V résineux')
    ws.write(9, 1, vol_res_tot)
    ws.write(10, 0, 'V Feuil. + res')
    ws.write(10, 1, vol_tot)
    
    #insérer image de plot préalablement convertie en png dans nouv fichier Excel
    ws.insert_image('A13', 'plot_résultats.png')
    
    ws.insert_image('K13', 'hist_class_diam.png')
    
    
    #fermer le nouveau fichier
    wb.close()
    print("=========================================================================")
    print("les résultats ont été enregistrés dans le nouveau fichier "
          "excel \"résultats inventaire.xlsx\"")
    print("=========================================================================")


entrer_path() #appels la première fonction


